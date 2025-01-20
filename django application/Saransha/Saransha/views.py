# my_app/views.py

from django.http import HttpResponse 
from django.shortcuts import render
from django.core.files.storage import FileSystemStorage
import pandas as pd
import openpyxl
import io
from .utils import load_and_filter_excel, get_publications_from_profile, process_profiles_from_excel, generate_author_summary,update_publication_details
from django.shortcuts import redirect
from django.http import JsonResponse
from graph_app.models import Users_Publication
from django.contrib import messages 
from . import scrap as s
from datetime import datetime
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl import Workbook
import os
#------------------------------------------------------------------------------------------------------------------------------------------
def upload_page(request):
    if "user_email" in request.session:
        excel_data = []
        publications_data = []
        if request.method == "POST" and request.FILES['excel_file']:

            excel_file = request.FILES['excel_file']
            fs = FileSystemStorage()
            filename = fs.save(excel_file.name, excel_file)
            file_path = fs.path(filename)

            output_file = fs.path("all_authors_publications.xlsx")
            process_profiles_from_excel(file_path, output_file)

            output_df = pd.read_excel(output_file)
            publications_data = output_df.to_dict(orient='records')

            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            for row in worksheet.iter_rows():
                row_data = [cell.value for cell in row]
                excel_data.append(row_data)

        return render(request, 'auth/upload.html', {'excel_data': excel_data, 'publications_data': publications_data})
    else:
        return upload_redirect(request)
        messages.warning(request, "Login required to access this page.")
        return render(request, 'upload.html')
        
     
    # return upload_redirect(request)

def upload_redirect(request):
    context = {'warning_message': "Login required to access this page."}
    return render(request, 'upload.html', context)
   #messages.warning(request, "Login required to access this page.")
   #return render(request, 'upload.html')
#------------------------------------------------------------------------------------------------------------------------------------------

def generatesummary(request):
    authors = []
    result_df = pd.DataFrame()
    summary= pd.DataFrame()
    faculty_member = ""
    start_year = None
    end_year = None
    sort_by = ""
    if "user_email" in request.session:
        if request.method == 'POST':
            
            faculty_member = request.POST.get('facultySelect', "")
            start_year = request.POST.get('startYear')
            end_year = request.POST.get('endYear')
            sort_by = request.POST.get('sortBy', "")

            try:
                start_year = int(start_year) if start_year else 0
                end_year = int(end_year) if end_year else 0
            except ValueError:
                start_year = 0
                end_year = 2025

        fs = FileSystemStorage()
        output_file_path = fs.path("all_authors_publications.xlsx")

        if fs.exists(output_file_path): 
            df = pd.read_excel(output_file_path)
            authors = df['Main Author'].unique().tolist()

            result_df = load_and_filter_excel(
                file_path=output_file_path,
                columns=['Main Author', 'Title','Journal','conference','Publication Type', "Year", "Cited by","co_author"],
                column_name='Main Author',
                valid_names=[faculty_member] if faculty_member != "All" else authors,
                year_range=[start_year , end_year],
                cited_by_sort_order= sort_by,
            )

            if "downloadSummary" in request.POST:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    result_df.to_excel(writer, index=False, sheet_name='Summary')
                buffer.seek(0)
                response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename="filtered_summary.xlsx"'

                return response
            
            if "generateSummary" in request.POST:
                result_df.to_excel('output.xlsx', index=False)
                data= generate_author_summary(result_df)
                data0 = data.to_html(classes='table table-striped', index=False)
                data0 = data0.replace('<tr style="text-align: right;">', '<tr style="text-align: left;">')

                data1=data.to_dict(orient='records')
            
                return render(request, "auth/generatesummary.html", {
            'authors': authors,
            'result_df': result_df.to_dict(orient='records'),
            'summary' :summary.to_dict(orient='records'),
            'data': data0,
            'data1': data1
        })  
                
                 
                return JsonResponse({'summary': summary.to_dict(orient='records')}, safe=False)

        return render(request, "auth/generatesummary.html", {
            'authors': authors,
            'result_df': result_df.to_dict(orient='records'),
            'summary' :summary.to_dict(orient='records')
        })   
    else:
        return render(request,"generatesummary.html")
#------------------------------------------------------------------------------------------------------------------------------
def home(request):
    if "user_email" in request.session:
        return render(request,'auth/index.html')
    else:
        return render(request,'index.html')

def settings(request):
    if "user_email" in request.session:
         return render(request,'auth/settings.html')
    else:
         return render(request,'settings.html')
   

def help(request):
    if "user_email" in request.session:
         return render(request,'auth/help.html')
    else:
         return render(request,'help.html')
        

def login(request):
    if request.method == "POST":
        uemail= request.POST['email']
        password = request.POST['password']
        USER=Users_Publication.objects.all().filter(user_email=uemail).first()
        if not USER:
            return HttpResponse("Invalid Email or Password")
        else:
            org_pass=USER.user_password
            if password==org_pass:
                # Set session data
                request.session["user_email"] = uemail

                # Add success message
                return render(request,'auth/index.html',{"m":"log in sucsessfully"}) # Redirect to the profile page
            else:
                return render(request, 'login.html', {'error': 'Invalid credentials. Please enter correct credentials.'})
    else:
        return render(request, 'login.html')



def signup(request):
    if(request.method == "POST"):
        uname = request.POST['username']
        uemail = request.POST['email']
        ucategory = request.POST['category']
        upassword = int(request.POST['password'])

        obj = Users_Publication(user_name=uname,user_email=uemail,user_password=upassword,user_category=ucategory)
        obj.save()
        messages.warning(request, "user Added")

        return redirect('login')
        #return  render(request,'login.html')
    else:
        return render(request,'signup.html',{'error': 'please fill the form'})
    
def logo_view(request):
    # Clear the session to log the user out
    request.session.flush()
    return  render(request,'index.html',{"m":"logout sucsessfully"}) 
    return render(request, 'index.html')

  
def cust_view(request):
    if request:
        pass

    return render(request, 'auth/cust.html')
import os

def missVal_view(request):
    authors = []
    filtered_data = []
    selected_author = request.GET.get('author', 'All')  # Default to 'All'
    selected_title = request.GET.get('title', None)  # Selected title
    prefill_data = {}  # Dictionary to hold prefilled values

    # Path to the Excel file
    fs = FileSystemStorage()
    output_file_path = fs.path("all_authors_publications.xlsx")

    try:
        # Load the Excel file into a DataFrame
        df = pd.read_excel(output_file_path)

        # Ensure the required columns exist in the DataFrame
        if 'Main Author' in df.columns and 'Title' in df.columns:
            authors = df['Main Author'].dropna().unique().tolist()

            # Filter data by author
            if selected_author != 'All':
                titles_df = df[df['Main Author'] == selected_author]
                if 'conference' in df.columns and 'Journal' in df.columns:
                    titles_df = titles_df[titles_df['conference'].isnull() & titles_df['Journal'].isnull()]
                filtered_data = titles_df['Title'].dropna().tolist()

            # Prefill data if a title is selected
            if selected_title:
                row = df[(df['Main Author'] == selected_author) & (df['Title'] == selected_title)]
                if not row.empty:
                    prefill_data = {
                        'journal_name': row.iloc[0].get('Journal', ''),
                        'conference_name': row.iloc[0].get('conference', ''),
                        'year': row.iloc[0].get('Year', '')
                    }

    except Exception as e:
        print(f"Error reading Excel file: {e}")

    # Handle form submission
    if request.method == 'POST':
        
        journal_name = request.POST.get('journalName', 'N/A')
        conference_name = request.POST.get('conferenceName', 'N/A')
        year = request.POST.get('year', 'N/A')
        if selected_author != "N/A" and selected_title != "N/A" and journal_name != "N/A" and conference_name != "N/A" and year != "N/A":
            update_publication_details(output_file_path, selected_author, selected_title, journal_name, conference_name, year)

        print(f"Selected Author: {selected_author}")
        print(f"Title: {selected_title}")
        print(f"Journal Name: {journal_name}")
        print(f"Conference Name: {conference_name}")
        print(f"Year: {year}")      
        print(update_publication_details)
        return render(request, 'auth/missVal.html', {
        'authors': authors,
        'Title': filtered_data,
        'selected_author': selected_author,
        'selected_title': selected_title,
        'prefill_data': prefill_data,
    })

    # Render the template with the required context
    return render(request, 'auth/missVal.html', {
        'authors': authors,
        'Title': filtered_data,
        'selected_author': selected_author,
        'selected_title': selected_title,
        'prefill_data': prefill_data,
    })




def publication_form(request):
    if request.method == 'POST':
        # Get form data
        main_author = request.POST.get('main_author')
        title = request.POST.get('title')
        journal = request.POST.get('journal', '')
        conference = request.POST.get('conference', '')
        year = request.POST.get('year')
        cited_by = request.POST.get('cited_by', 0)
        coauther="N/A"
        Last_Search_Date=datetime.now().strftime("%Y-%m-%d")

        # File path to save Excel (can be dynamic as needed)
        file_path = os.path.join('media', 'publications.xlsx')
            # Path to the Excel file
        fs = FileSystemStorage()
        output_file_path = fs.path("all_authors_publications.xlsx")

        # Check if the file already exists
        if os.path.exists(output_file_path):
            from openpyxl import load_workbook
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            # Create a new workbook
            workbook = Workbook()
            sheet = workbook.active
            # Add header row
            sheet.append(["Main Author", "Title", "Journal", "Conference", "Year", "Cited By"])

        # Append data
        sheet.append([main_author, title, journal, conference, year, cited_by])

        # Save the file
        workbook.save(output_file_path)

        return HttpResponse("Publication saved successfully!")

    return render(request, 'publication_form.html')
