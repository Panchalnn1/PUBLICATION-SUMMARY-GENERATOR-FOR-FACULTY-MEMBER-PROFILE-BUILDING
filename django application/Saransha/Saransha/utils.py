# my_app/utils.py

import os
import pandas as pd
import io
from scholarly import scholarly
from datetime import datetime

import requests


# Scopus API details
SCOPUS_URL = "https://api.elsevier.com/content/serial/title"

# Function to check if a journal is indexed in Scopus and return the updated DataFrame
def check_scopus_index_for_df(df, api_key):
    """
    Takes a DataFrame with a 'Journal' column, checks whether each journal is indexed in Scopus,
    and returns the updated DataFrame with a new 'Scopus Indexed' column.

    Parameters:
    df (pd.DataFrame): Input DataFrame with a 'Journal' column containing journal names.
    api_key (str): Your Scopus API key.

    Returns:
    pd.DataFrame: Updated DataFrame with a 'Scopus Indexed' column.
    """
    SCOPUS_URL = "https://api.elsevier.com/content/serial/title"
    
    # Function to initialize API headers with your API key
    def get_headers():
        return {
            'Accept': 'application/json',
            'X-ELS-APIKey': api_key,
        }
    
    # Function to check if a journal is indexed in Scopus
    def check_scopus_index(journal_name):
        headers = get_headers()
        params = {'title': journal_name}
        response = requests.get(SCOPUS_URL, headers=headers, params=params)
        
        if response.status_code == 200:
            data = response.json()
            if data.get('serial-metadata-response'):
                return True  # Journal is indexed
            return False  # Journal not indexed
        else:
            return None  # Error occurred (invalid response)

    # Ensure 'Journal' column exists in the DataFrame
    if 'Journal' not in df.columns:
        raise ValueError("The input DataFrame must have a 'Journal' column.")
    
    # Check Scopus indexing status for each journal and add it as a new column
    df['Scopus Indexed'] = df['Journal'].apply(lambda journal: check_scopus_index(journal))
    
    return df



def load_and_filter_excel(file_path, sheet_name='Sheet1', columns=None, column_name=None, valid_names=None, cited_by_sort_order=None, year_range=None):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"The file {file_path} does not exist")

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
    except Exception as e:
        raise Exception(f"Failed to load the Excel file: {str(e)}")

    if columns:
        missing_columns = [col for col in columns if col not in df.columns]
        if missing_columns:
            print(f"Warning: The following columns are not in the sheet: {', '.join(missing_columns)}")
            columns = [col for col in columns if col in df.columns]
        df = df[columns]

    if column_name and valid_names:
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' does not exist in the DataFrame")
        df = df[df[column_name].isin(valid_names)]

    if cited_by_sort_order:
            if cited_by_sort_order == 'Date':
                if 'Year' not in df.columns:
                    raise ValueError("The column 'Year' does not exist in the DataFrame")
                df = df.sort_values(by='Year', ascending=True)  # Sort by Year in ascending order
            elif cited_by_sort_order in ['asc', 'desc'] and 'Cited by' in df.columns:
                df['Cited by'] = pd.to_numeric(df['Cited by'], errors='coerce').fillna(0).astype(int)
                df = df.sort_values(by='Cited by', ascending=(cited_by_sort_order == 'asc'))
        

    if year_range and len(year_range) == 2:
        if 'Year' in df.columns:
            df = df[(df['Year'] >= year_range[0]) & (df['Year'] <= year_range[1])]
       
        

    return df

def get_publications_from_profile(profile_url):
    user_id = profile_url.split("user=")[1].split("&")[0]
    author = scholarly.search_author_id(user_id)
    scholarly.fill(author)
    main_author = author['name']
    current_date = datetime.now().strftime("%Y-%m-%d")

    publications_data = []

    for pub in author['publications']:
        scholarly.fill(pub)
        publication_details = {
            'Main Author': main_author,
            'Title': pub['bib']['title'],
            'conference': pub['bib'].get('conference', 'N/A'),
            'Journal': pub['bib'].get('journal', 'N/A'),
            'Year': pub['bib'].get('pub_year', 'N/A'),
            'Publication Type': pub['bib'].get('ENTRYTYPE', 'N/A'),
            'Cited by': pub.get('num_citations', 'N/A'),
            "co_author":pub['bib']['author'],
            'Last Search Date': current_date
        }
        publications_data.append(publication_details)

    return publications_data

def process_profiles_from_excel(file_path, output_file):
    profiles_df = pd.read_excel(file_path)
    all_publications = []

    for profile_url in profiles_df['Profile URL']:
        try:
            publications = get_publications_from_profile(profile_url)
            all_publications.extend(publications)
        except Exception as e:
            print(f"Could not retrieve publications for {profile_url}: {e}")

    publications_df  = pd.DataFrame(all_publications)
    #API_KEY = 'd93d31bed1f4166cb5cda30e1718ea5c'
   # publications_df = check_scopus_index_for_df(publications_df, API_KEY)

    


    publications_df .to_excel(output_file, index=False)
    print(f"All publication details saved to {output_file}")


def generate_author_summary(df):
    df.columns = df.columns.str.strip()
    summary = df.groupby('Main Author').agg(
        publication=('Title', 'count'),
        journal=('Journal', lambda x: x.notnull().sum()),
        total_citations=('Cited by', 'sum')
        ).reset_index()
    
    summary = summary.sort_values(by='total_citations', ascending=False)
    
  
    return summary

import pandas as pd

def generate_publication_summary(dataframe):
    # Define the range of years
    years = list(range(min(dataframe['Year']),max(dataframe['Year'])))

    # Get a list of unique authors
    authors = dataframe['Main Author'].unique()

    # Create an empty dictionary to store publication counts
    publications = {author: [0] * len(years) for author in authors}

    # Count publications for each author by year
    for _, row in dataframe.iterrows():
        author = row['Main Author']
        year = row['Year']
        if year in years:  # Ensure the year is within the range
            year_index = years.index(year)
            publications[author][year_index] += 1

    # Convert years to strings for the final output
    years = list(map(str, years))

    return years, publications


from openpyxl import load_workbook

def update_publication_details(file_path, author, title, new_journal_name, new_conference_name, new_year):
    # Load the workbook
    try:
        workbook = load_workbook(file_path)
    except Exception:
        return "Failed to load the Excel file."

    # Select the active sheet or specify the sheet name
    sheet = workbook.active  # Or use sheet = workbook['SheetName'] to select a specific sheet
    
    # Loop through rows and columns to access and find the matching row
    for row in sheet.iter_rows():
        if row[0].value == author and row[1].value == title:
            # Modify the values in the found row
            row[2].value = new_journal_name  # Example: Update Journal Name in the 3rd column
            row[3].value = new_conference_name  # Example: Update Conference Name in the 4th column
            row[4].value = new_year  # Example: Update Year in the 5th column (adjust the index if necessary)

            # Save the workbook after modification
            try:
                workbook.save(file_path)  # Save changes to the same file
                return "Update successful."
            except Exception:
                return "Failed to save the Excel file."
    
    # If no matching row was found
    return "Row with the given author and title not found."
